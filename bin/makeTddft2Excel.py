import numpy as np
import matplotlib.pyplot as plt
import pyscf
from pyscf.geomopt.geometric_solver import optimize
from pyscf import gto, scf, dft, tddft
from rdkit import Chem
from rdkit.Chem import rdDepictor
from rdkit.Chem.Draw import rdMolDraw2D
from rdkit.Chem.Draw.MolDrawing import DrawingOptions

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU, pixels_to_points
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.drawing.text import CharacterProperties
import openpyxl.drawing.image

import subprocess
import s00Utils
import math,os,sys
from io import BytesIO
from cairosvg import svg2png
import PIL

def gaussBand(x, band, strength, stdev):
  constant=0.0003
  bandshape = constant * (strength/(1.0/stdev))*np.exp(-(((1.0/x)-(1.0/band))**2/(1.0/stdev)**2))
  return bandshape

def peak(xs):
    m=0
    for x in xs:
        if m<x:
            m=x
    return f':{m:.2f}'

def norm(xs):
    ret=[]
    m=0
    for x in xs:
        if m<x:
            m=x

    for x in xs:
        ret.append(x/m)
    return ret

def getOpenRow(xs,ys,ws):
    while True:
        if ws.cell(column=xs,row=ys).value==None:
            return ys
        print(ws.cell(column=xs,row=ys).value)
        ys=ys+1

def generate_image(mol, size):

    image_data = BytesIO()
    view = rdMolDraw2D.MolDraw2DSVG(size[0], size[1])
    tm = rdMolDraw2D.PrepareMolForDrawing(mol)

    view.SetLineWidth(1)
    view.SetFontSize(1.3 * view.FontSize())
    option = view.drawOptions()    
    option.multipleBondOffset = 0.09
    option.padding = 0.11
    view.DrawMolecule(tm)
    view.FinishDrawing()
    svg = view.GetDrawingText()
    svg2png(bytestring=svg, write_to=image_data)
    #image_data.seek(0)
    return PIL.Image.open(image_data)

def set_sell_image(ws, row, col, mol, size=(300, 150), margin=5, rate=1.0):
    # https://stackoverflow.com/questions/55309671/more-precise-image-placement-possible-with-openpyxl-pixel-coordinates-instead
    image = generate_image(mol, (int(size[0]*rate), int(size[1]*rate)))
    image =  openpyxl.drawing.image.Image(image)
    col_offset = pixels_to_EMU(margin)
    row_offset = pixels_to_EMU(margin)
    size_ext = XDRPositiveSize2D(pixels_to_EMU(size[0]), pixels_to_EMU(size[1]))
    maker = AnchorMarker(col=col, colOff=col_offset, row=row-1, rowOff=row_offset)
    image.anchor = OneCellAnchor(_from=maker, ext=size_ext)
    ws.add_image(image)

###############################
###### main
###############################
if len(sys.argv)==2 and sys.argv[1]=='-h':
    print("version 2.0")
    print(sys.argv[0],"-d input_path | -n tag_name -k key (key*.tdd only) -o output.excel -range 100,600,501")
    print("summary without measured data")
    exit()

key=False
path=".";skip=False;name='tddft';

#output='test.xlsx'
if os.environ['NAME']=='kisa-pc':
    k='uv'
else:
    k='wsl'

try:
    wk=os.getcwd().split(k)[1].split('/')
except:
    wk=['uv']
output="_".join(wk)+"_all.xlsx"

x_range=[100,600,501];stdev=4000
for n,arg in enumerate(sys.argv):
    if skip:
        skip=False
        continue
    if arg =='-d':
        path=sys.argv[n+1]
        skip=True
    elif arg =='-n':
        name=sys.argv[n+1]
        skip=True
    elif arg =='-k':
        key=sys.argv[n+1]
        skip=True
    elif arg =='-range':
        x_range=sys.argv[n+1].split(',')
    elif arg =='-stdev':
        stdev=float(sys.argv[n+1])
    elif arg =='-o':
        output=sys.argv[n+1]
        skip=True

fns=[]
files=os.listdir(path)
for f in files:
    if f.split(".")[-1]=='tdd':
#        print(f)
        if key!=False:
            if f.startswith(key):
                fns.append(f.split(".")[0])
        else:
            fns.append(f)
fns.sort()
print("-->",fns)

if not os.path.isfile(output):
    wb = openpyxl.Workbook()
    wb.save(output)

wb = openpyxl.load_workbook(output)

# b2: 製品名、c:SMILES、d:UVスペクトル：
sheets=wb.sheetnames

summary="一覧";thisList=name

item=['製品名','SMILES','構造式','UVスペクトル','Gap','HOMO','LUMO','λpeak']
if not summary in sheets:
    wb.create_sheet(title=summary)
    ws=wb[summary]
    for i in range(len(item)):
        ws.cell(row=2,column=2+i,value=item[i])
        ws[chr(66+i)+'2'].font=Font(size=40)

for ws in wb.worksheets:
    if ws.title.startswith('Sheet'):
        wb.remove(ws)

ws=wb[summary]

width=[]
for i in range(5):
    width.append(ws.column_dimensions[chr(i+65)].width)

wb.create_sheet(title=thisList)
wd=wb[thisList]

x=np.linspace(int(x_range[0]),int(x_range[1]),int(x_range[2]))

col=1
for i,xx in enumerate(x):
    wd.cell(row=2+i,column=col,value=xx)
xval=openpyxl.chart.Reference(wd,min_col=1,min_row=2,max_row=2+len(x))

ys=yso=getOpenRow(2,2,ws)
col=col+1
for fn in fns:
    product_name=fn.split('.')[0]
    ws.cell(row=ys,column=2,value=product_name) # yes product name
    ws['B'+str(ys)].font=Font(size=40)
    ws.cell(row=ys,column=2).alignment=Alignment(vertical='center')
    ws.cell(row=ys,column=3).alignment=Alignment(vertical='center')
    ws.cell(row=ys,column=4).alignment=Alignment(horizontal='center',vertical='center')
    ww=(len(product_name)+2)*40/12
    if width[1]<ww:
        width[1]=ww
    w,f,s,h=s00Utils.getNP(path+"/"+fn.split(".tdd")[0])
    if len(w)<1:
        continue
    ws.cell(row=ys,column=3,value=s) # yes SMILES
    ws['C'+str(ys)].font=Font(size=40)
    ww=(len(s)+2)*40/12
    if width[2]<ww:
        width[2]=ww
    cmp=s00Utils.getCP(x,w,f,stdev)
    peaks=s00Utils.getPeaks(x,cmp,xlim=[100,400],exel=True)
    wd.cell(row=1,column=col,value=product_name)
    for i,xx in enumerate(cmp):
        wd.cell(row=i+2,column=col,value=xx)

    ws.row_dimensions[ys].height=200
    set_sell_image(ws,ys,3,Chem.MolFromSmiles(s),size=(200,200))
    ws.cell(row=ys,column=2).alignment=Alignment(vertical='center')

    yval=openpyxl.chart.Reference(wd,min_col=col,min_row=2,max_row=2+len(x))

    chart = openpyxl.chart.ScatterChart('line')
    chart.x_axis.numFmt = "000"
    chart.y_axis.numFmt = "0.0"
#    chart.x_axis.title='wavelength (nm)'
#    chart.y_axis.title='Absorption'

    series=openpyxl.chart.Series(yval,xval)
    chart.series.append(series)
    chart.series[0].graphicalProperties.line.width=100
    chart.legend=None
    chart.width=12;chart.height=6
#    chart.width=6;chart.height=3
    chart.x_axis.scaling.min=int(x_range[0])
    chart.x_axis.scaling.max=int(x_range[1])
    chart.style=2

    ws.add_chart(chart,'E'+str(ys))

#    putCell(tbl,i+1,1,"{:.4f}".format(l[0]))
#                 putCell(tbl,i+1,2,"{:.4f}".format(l[1]))
#                putCell(tbl,i+1,3,"{:.4f}".format(l[1]-l[0]))
    if h[0]!=False:
        ws.cell(row=ys,column=6,value=h[1]-h[0])
        ws.cell(row=ys,column=7,value=h[0])
        ws.cell(row=ys,column=8,value=h[1])
        ws['F'+str(ys)].font=Font(size=40)
        ws['G'+str(ys)].font=Font(size=40)
        ws['H'+str(ys)].font=Font(size=40)

    for i,p in enumerate(peaks):
        ws.cell(row=ys,column=9+i,value=p)
        ws[chr(i+73)+str(ys)].font=Font(size=40)
#        ws.column_dimensions[chr(i+73)].width=30

    col=col+1

    ys=ys+1

width[3]=30;width[4]=70
print(width[2])
for i in range(5):
    ws.column_dimensions[chr(i+65)].width=width[i]
for i in range(7):
    ws.column_dimensions[chr(i+70)].width=30

ws.sheet_view.zoomScale=30
ws.sheet_view.zoomScaleNormal=30
wb.save(output)

exit()
